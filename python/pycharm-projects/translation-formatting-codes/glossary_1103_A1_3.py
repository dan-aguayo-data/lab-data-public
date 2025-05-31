from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Define the directory path and date
save_directory = r"C:\Users\DanielAguayo\OneDrive - Circular Economy Systems\Desktop"
current_date = "2025-05-11"

# Function to save glossary with color coding and instructions
def save_glossary_with_color(data, headers, filename, key_terms, has_remarks=False, editor="DanielAguayo"):
    wb = Workbook()
    ws = wb.active

    # Add instructions at the top
    ws.append(["Instructions", ""])
    ws.append(["Color Coding", "Green (hex: 90EE90) for key terms; Yellow (hex: FFFF99) for remarks"])
    ws.append(["Version Control", "Indicated in filename (e.g., v1_DanielAguayo)"])
    ws.append(["Sequence Convention", "Alphabetical order for easy reference"])
    ws.append(["", ""])  # Spacer row

    # Write headers
    ws.append(headers)

    # Write data
    for row_data in data:
        ws.append(row_data)

    # Apply color coding
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row in range(7, ws.max_row + 1):  # Start after instructions (row 7)
        term = ws[f"A{row}"].value
        remarks = ws[f"E{row}"].value if has_remarks else None
        if term in key_terms:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill
        if has_remarks and remarks:
            ws.cell(row=row, column=5).fill = yellow_fill

    # Add editing authority note
    ws.append(["", ""])  # Spacer row
    ws.append(["Editing Authority", f"Editable only by {editor}"])

    # Save the workbook
    full_path = f"{save_directory}\\{filename}"
    wb.save(full_path)
    print(f"Saved: {full_path}")

# Glossary: Bilingual Glossary for ATE23
glossary_headers = ["Term (English)", "Translation (Spanish)", "Synonym", "Source", "Remarks"]
glossary_data = [
    ["Amendment Charges", "Cargos por modificaciones", "Change Fees", "Tourism Australia", "Fees for booking changes"],
    ["Australian Tourism Exchange (ATE)", "Intercambio Turístico Australiano (ATE)", "Tourism Trade Event", "Tourism Australia", "Main event title"],
    ["Booth", "Stand", "Exhibit Space", "OnFulfillment", "Space for exhibitors at ATE23"],
    ["Business-to-Business (B2B)", "Empresa a empresa (B2B)", "B2B Model", "Tourism Australia", "Event format"],
    ["Canopy", "Dosel", "Booth Covering", "OnFulfillment", "Used for branding at booths"],
    ["Coastal Landscape", "Paisaje costero", "Beach Setting", "Tourism Australia", "Event theme"],
    ["Cultural Tourism", "Turismo cultural", "Heritage Tourism", "UNWTO", "Type of tourism at ATE23"],
    ["Destination", "Destino", "Travel Location", "UNWTO", "e.g., Gold Coast (Costa Dorada)"],
    ["Ecotourism", "Ecoturismo", "Nature-Based Tourism", "UNWTO", "Focus on minimal impact"],
    ["Exhibitor", "Expositor", "Vendor", "OnFulfillment", "Displays at ATE23"],
    ["Foyer", "Vestíbulo", "Lobby", "OnFulfillment", "Entrance area for networking"],
    ["Familiarisation (Famil) Program", "Programa de familiarización (famil)", "Introductory Trip", "Tourism Australia", "Pre/post-event activity"],
    ["Gold Coast", "Costa Dorada", "Gold Coast Region", "Tourism Australia", "Event location"],
    ["Group Conditions", "Condiciones para grupos", "Group Terms", "Tourism Australia", "Terms for group bookings"],
    ["GST", "Impuesto sobre bienes y servicios (GST)", "Goods and Services Tax", "Tourism Australia", "10% tax in Australia"],
    ["Inbound Tour Operator", "Operador turístico de entrada", "Travel Organizer", "Tourism Australia", "Arranges travel for visitors"],
    ["International Buyer", "Comprador internacional", "Global Buyer", "Tourism Australia", "Event delegate"],
    ["Networking Event", "Evento de networking", "Business Networking", "Tourism Australia", "Key event component"],
    ["Outback Landscape", "Paisaje del interior", "Desert Setting", "Tourism Australia", "Event theme"],
    ["Rainforest Landscape", "Paisaje de selva tropical", "Jungle Setting", "Tourism Australia", "Event theme"],
    ["Reconciliation Australia Acknowledgement", "Reconciliación Australia reconoce a los Propietarios Tradicionales...", "Traditional Owners Acknowledgement", "Reconciliation Australia", "Formal event statement"],
    ["Seller Delegate", "Delegado vendedor", "Australian Seller", "Tourism Australia", "Event delegate"],
    ["Sponsorship Package", "Paquete de patrocinio", "Sponsorship Deal", "OnFulfillment", "Benefits for sponsors"],
    ["Sustainable Tourism", "Turismo sostenible", "Eco-Tourism", "UNWTO", "Focus on minimal impact"],
    ["Tourism Product", "Producto turístico", "Tourism Offering", "UNWTO", "Attractions and services"]
]

# File details
filename = f"ATE23GoldCoast_Glossary_v1_DanielAguayo.xlsx"
key_terms = ["Australian Tourism Exchange (ATE)", "Trade Show", "Sustainable Tourism", "Reconciliation Australia Acknowledgement"]

# Save the glossary
save_glossary_with_color(glossary_data, glossary_headers, filename, key_terms, has_remarks=True)

print("Glossary created successfully!")