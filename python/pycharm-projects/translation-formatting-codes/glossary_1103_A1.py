from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Define the directory path and date
save_directory = r"C:\Users\DanielAguayo\OneDrive - Circular Economy Systems\Desktop"
current_date = "2025-05-11"

# Function to save glossary with color coding and instructions
def save_glossary_with_color(data, headers, filename, key_terms, editor="DanielAguayo"):
    wb = Workbook()
    ws = wb.active

    # Add instructions at the top
    ws.append(["Instructions", ""])
    ws.append(["Color Coding", "Green (hex: 90EE90) for key terms; Yellow (hex: FFFF99) for remarks"])
    ws.append(["Version Control", "Indicated in filename (e.g., v1_DanielAguayo); Modified By column tracks edits"])
    ws.append(["Sequence Convention", "Alphabetical order for easy reference"])
    ws.append(["", ""])  # Spacer row

    # Write headers
    ws.append(headers)

    # Write data
    for row_data in data:
        ws.append(row_data)

    # Apply color coding
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    for row in range(7, ws.max_row + 1):  # Start after instructions (row 7)
        term = ws[f"A{row}"].value
        if term in key_terms:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

    # Add editing authority note
    ws.append(["", ""])  # Spacer row
    ws.append(["Editing Authority", f"Editable only by {editor}"])

    # Save the workbook
    full_path = f"{save_directory}\\{filename}"
    wb.save(full_path)
    print(f"Saved: {full_path}")

# Glossary: Updated Strata Plan Glossary with Spanish Translations
glossary_headers = ["English", "Definition", "LOTE (Spanish Translation)", "Source (Translation)", "Modified By"]
glossary_data = [
    ["By-laws", "These are a set of rules that govern a range of matters.", "Estatutos", "IATE", "Daniel Aguayo"],
    ["Common property", "This term refers to the land in a strata titles scheme which is owned jointly by all of the lot owners.", "Propiedad común", "IATE", "Daniel Aguayo"],
    ["Lots (referring to parts of the parcel)", "These are parts of the parcel of land under a strata titles scheme, each of which has its own title and can be owned and dealt with separately from the other lots in the scheme.", "Lotes", "SpanishDict", "Daniel Aguayo"],
    ["Parcel", "This is an area of land under a strata titles scheme.", "Parcela", "SpanishDict", "Daniel Aguayo"],
    ["Resolutions (of a strata company)", "These are the decisions of a strata company which are made by taking the votes of members.", "Resoluciones (de una compañía de estrato)", "IATE, SpanishDict", "Daniel Aguayo"],
    ["Strata title scheme", "This is a kind of property ownership by which an area of land can be divided into two or more lots and common property.", "Esquema de título de estrato", "IATE, SpanishDict", "Daniel Aguayo"],
    ["Strata Titles Act 1985 (WA)", "This is an Act regulating strata titles in Western Australia, passed by the WA parliament in 1985.", "Ley de Títulos de Estrato 1985 (WA)", "IATE", "Daniel Aguayo"],
    ["Survey strata scheme", "This is one kind of strata titles scheme.", "Esquema de estrato topográfico", "IATE, SpanishDict", "Daniel Aguayo"],
    ["Unit entitlements", "These figures represent the proportionate value of each owner’s share in a strata titles scheme.", "Derechos de unidad", "IATE", "Daniel Aguayo"]
]

# File details
filename = f"StrataPlanGlossary_v1_DanielAguayo.xlsx"
key_terms = ["Strata Titles Act 1985 (WA)", "Strata title scheme", "Common property"]

# Save the glossary
save_glossary_with_color(glossary_data, glossary_headers, filename, key_terms)

print("Glossary created successfully!")