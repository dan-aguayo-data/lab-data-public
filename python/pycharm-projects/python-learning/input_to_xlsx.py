from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Define the directory path and date
save_directory = r"D:\Studies\Babel\PSPTIS103"
current_date = "2025-04-05"


# Function to save glossary with color coding
def save_glossary_with_color(data, headers, filename, key_terms, has_remarks=False):
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = wb.active

    # Write headers
    ws.append(headers)

    # Write data
    for row_data in data:
        ws.append(row_data)

    # Define color fills
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Light green for key terms
    yellow_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # Light yellow for remarks

    # Apply color coding
    for row in range(2, ws.max_row + 1):  # Start from row 2 (after header)
        term = ws[f"A{row}"].value  # Term is in column A
        remarks = ws[f"E{row}"].value if has_remarks else None

        # Color key terms (green)
        if term in key_terms:
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = green_fill

        # Color remarks (yellow) if applicable
        if has_remarks and remarks:
            ws.cell(row=row, column=5).fill = yellow_fill  # Column E for Remarks

    # Save the workbook
    full_path = f"{save_directory}\\{filename}"
    wb.save(full_path)
    print(f"Saved: {full_path}")


# Glossary 1: Established Glossary (Monolingual)
glossary1_headers = ["Term", "Definition", "Source", "Link"]
glossary1_data = [
    ["Assault",
     "The direct or indirect application of force to another person, or the attempt or threat to apply force, without consent.",
     "Toronto Police Service Glossary", "https://www.tps.ca"],
    ["Common Assault", "An assault causing minimal harm, often involving threats or minor physical contact.",
     "NSW Crime Tool Glossary", "https://www.bocsar.nsw.gov.au"],
    ["Offence",
     "A violation against the Criminal Code or other statute, reported in the year it comes to police attention.",
     "Toronto Police Service Glossary", "https://www.tps.ca"],
    ["Victim", "Any person who has been subjected to an offence, such as assault.", "Toronto Police Service Glossary",
     "https://www.tps.ca"],
    ["Investigation Status",
     "Indicates how an offence has been dealt with by police, such as 'unsolved' or 'charges laid,' at the time of data extraction.",
     "NSW Crime Tool Glossary", "https://www.bocsar.nsw.gov.au"],
    ["Charge", "A formal accusation laid by police against a suspect for an offence.", "NSW Crime Tool Glossary",
     "https://www.bocsar.nsw.gov.au"],
    ["Interview", "A formal process where police question a suspect, victim, or witness to gather evidence.",
     "Western Australia Police Force Glossary", "https://www.police.wa.gov.au"],
    ["Arrest", "Taking a suspect into custody for an alleged offence.", "Toronto Police Service Glossary",
     "https://www.tps.ca"],
    ["Bail", "The release of a suspect with conditions pending court appearance.", "NSW Crime Tool Glossary",
     "https://www.bocsar.nsw.gov.au"],
    ["Evidence", "Information or items collected to support the investigation of an offence.",
     "Western Australia Police Force Glossary", "https://www.police.wa.gov.au"]
]
filename1 = f"CommonAssaultPoliceInterview_Glossary_1_{current_date}.xlsx"
key_terms1 = ["Assault", "Common Assault", "Offence"]
save_glossary_with_color(glossary1_data, glossary1_headers, filename1, key_terms1, has_remarks=False)

# Glossary 2: Self-Developed Bilingual Glossary
glossary2_headers = ["Term (English)", "Translation (Spanish)", "Synonym", "Source", "Remarks"]
glossary2_data = [
    ["Assault", "Asalto", "Attack", "Oxford Dictionary", "Core offence in interview"],
    ["Common Assault", "Asalto común", "Minor Assault", "NSW Crime Tool Glossary", "Involves minimal harm"],
    ["Offence", "Delito", "Crime", "Toronto Police Service Glossary", "Violation of law"],
    ["Victim", "Víctima", "Complainant", "Toronto Police Service Glossary", "Person harmed"],
    ["Suspect", "Sospechoso", "Accused", "Oxford Dictionary", "Person under investigation"],
    ["Interview", "Entrevista", "Questioning", "Western Australia Police Force Glossary", "Police questioning session"],
    ["Charge", "Cargo", "Accusation", "NSW Crime Tool Glossary", "Formal legal action"],
    ["Arrest", "Arresto", "Detention", "Toronto Police Service Glossary", "Taking into custody"],
    ["Bail", "Fianza", "Release", "NSW Crime Tool Glossary", "Conditional release"],
    ["Evidence", "Evidencia", "Proof", "Western Australia Police Force Glossary", "Supports case"],
    ["Statement", "Declaración", "Testimony", "Oxford Dictionary", "Recorded account"],
    ["Witness", "Testigo", "Observer", "SpanishDict", "Provides testimony"],
    ["Investigation", "Investigación", "Inquiry", "NSW Crime Tool Glossary", "Police process"],
    ["Consent", "Consentimiento", "Agreement", "Toronto Police Service Glossary", "Relevant to assault cases"],
    ["Custody", "Custodia", "Detention", "Oxford Dictionary", "Holding at station"]
]
filename2 = "CommonAssaultPoliceInterview_Glossary_2_v1_DanielAguayo.xlsx"
key_terms2 = ["Assault", "Common Assault", "Offence"]
save_glossary_with_color(glossary2_data, glossary2_headers, filename2, key_terms2, has_remarks=True)

print("Glossaries created successfully!")